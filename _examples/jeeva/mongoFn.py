#dependency are pymongo,openpyxl
#pip install pymongo
#pip install openpyxl
#install mongodb and configure to this port no 27017 

import pymongo
import csv
from openpyxl import load_workbook
from pprint import pprint

client = pymongo.MongoClient("localhost",27017)
db = client["back_testing"]
excel = db["excel"]
#db name "back testing"
#collection or table name "excel"

def dbInsert():
    json={
        "location":"C:/Users/Jeevanatham N/Documents/csv&xlsx/sample.csv",
        "filename":"sample",
        "type":"csv"
    }
    #the json inserted into collection
    excel.insert_one(json)

def getFile():
    files = excel.find({"filename":"sample"})
    #finding values by using keys
    #may be files object have multiple results so i iterated
    for file in files:
        #if filetype is excel
        if (file["type"] == "xlsx"):
            #i just excelParse function
            excelParse(file["location"],file["filename"])
        #if filetype is csv 
        elif (file["type"] == "csv"):
            #i just csvParse function
            csvParse(file["location"],file["filename"])

def excelParse(filepath,filename):
    #this method parse excel into array like this format [{},{},{}]
    excelValue = []
    json = {}
    #the filepath argument have excel file location using that location i opening excel file "load_workbook"
    wb=load_workbook(filepath)
    #may be it have multiple sheets i selecting currently active sheet
    sheet=wb.active
    max_row=sheet.max_row
    max_column=sheet.max_column
    for i in range(2,max_row+1):
        json = {}
        for j in range(1,max_column+1):
            cell_obj=sheet.cell(row=i,column=j)
            cell_title=sheet.cell(row=1,column=j)
            json[cell_title.value] = cell_obj.value
        excelValue.append(json)
    #by iterating we get all excel values in the form of [{},{},{},{}]

def csvParse(filepath,filename):
    #this method parse csv into array like this format [{},{},{}]
    with open(filepath, mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        csv_reader = list(csv_reader)
    #we get all csv values in the form of [{},{},{},{}]
    overrideCsv(filepath,filename,csv_reader)

def overrideCsv(filepath,filename,csv_reader):
    #this method after crud operation when user hit save button response come here 
    #response object should be like this array of json [{},{},{}] "harish nee indha format la send pannu"  
    with open(filepath, mode='w') as csv_file:
        fieldnames = []
        for i in csv_reader:
            print(i)
        for key in csv_reader[0]:
            fieldnames.append(key)
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        for i in csv_reader:
            writer.writerow(i)
    #after these operation new data will be updated guys!

def overrideExcel():
    #this method after crud operation when user hit save button response come here 
    #response object should be like this array of json [{},{},{}] "harish nee indha format la send pannu"
    #code no yet did
    pass

if __name__ == "__main__":
    #call either dbInsert() or getFile()
    getFile()
