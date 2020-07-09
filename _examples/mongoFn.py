import pymongo
import openpyxl
from openpyxl import load_workbook
from pprint import pprint

client = pymongo.MongoClient("localhost",27017)
db = client["back_testing"]
excel = db["excel"]

def dbInsert():
    json={
        "location":"C:/Users/Jeevanatham N/Downloads/report.xlsx",
        "name":"report"
    }
    excel.insert_one(json)

def getExcel():
    files = excel.find({"name":"sample"})
    for file in files:
        excelParse(file["location"],file["name"])

def excelParse(filepath,filename):
    excelValue = []
    row = []
    wb=load_workbook(filepath)
    sheet=wb.active
    max_row=sheet.max_row
    max_column=sheet.max_column
    for i in range(1,max_column+1):
        cell_obj=sheet.cell(row=1,column=i)
        row.append(cell_obj.value)
    excelValue.append(row)
    for i in range(2,max_row+1):
        row = []
        for j in range(2,max_column+1):
            cell_obj=sheet.cell(row=i,column=j)
            row.append(cell_obj.value)
        excelValue.append(row)
    for row in excelValue:
        print(row)

if __name__ == "__main__":
    getExcel()
